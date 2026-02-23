import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from openai import OpenAI
import os
from PIL import Image

# Custom CSS to set the font to Yu Gothic
st.markdown("""
<style>
.st-emotion-cache-10trblm, .st-emotion-cache-1629p8f, .st-emotion-cache-1c7y2kd {
    font-family: 'Yu Gothic', 'YuGothic', 'Hiragino Kaku Gothic ProN', 'Hiragino Sans', Meiryo, sans-serif;
}
</style>
""", unsafe_allow_html=True)

# 表題を画像で表示（拡大版）
try:
    header_image = Image.open("/home/ubuntu/streamlit_app/header_image.png")
    st.image(header_image, width=600)
except FileNotFoundError:
    st.title("専門的支援実施計画書 自動ツール")

st.write("Excelファイルをアップロードするだけで、AIが内容を自動更新し、ダウンロードできます。")

# OpenAI APIキーの設定
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    st.error("OpenAI APIキーが設定されていません。環境変数 OPENAI_API_KEY を設定してください。")
    st.stop()

client = OpenAI(api_key=OPENAI_API_KEY)

# パスワード認証
correct_password = "hanagurisekine"

if "password_entered" not in st.session_state:
    st.session_state["password_entered"] = False

if not st.session_state["password_entered"]:
    password = st.text_input("パスワードを入力してください", type="password")
    if st.button("ログイン"):
        if password == correct_password:
            st.session_state["password_entered"] = True
            st.rerun()
        else:
            st.error("パスワードが間違っています。")
else:
    # Excelファイルアップロードエリア
    uploaded_file = st.file_uploader("テンプレートExcelをアップロード", type=["xlsx"])

    if uploaded_file:
        st.info("Excelファイルを読み込み、AIが内容を生成・更新中です...少々お待ちください。")
        
        try:
            # BytesIOオブジェクトをリセットして、再度読み込めるようにする
            uploaded_file.seek(0)
            wb = load_workbook(uploaded_file)
            ws = wb.active

            # V33の総括をB12のアセスメントに転記し、V33を空欄にする
            if ws['V33'].value:
                ws['B12'] = ws['V33'].value
                ws['V33'] = None # V33を空欄にする
                st.info("V33の総括をB12のアセスメントに転記し、V33を空欄にしました。")

            # 既存のExcelファイルから情報を読み取る
            excel_data = []
            if ws['B12'].value:
                excel_data.append(f"既存のアセスメント: {ws['B12'].value}")
            if ws['P12'].value:
                excel_data.append(f"既存のスモールステップ: {ws['P12'].value}")
            if ws['L19'].value:
                excel_data.append(f"既存の長期目標: {ws['L19'].value}")
            if ws['L21'].value:
                excel_data.append(f"既存の短期目標: {ws['L21'].value}")
            if ws['C25'].value:
                excel_data.append(f"既存の具体的支援内容1: {ws['C25'].value}")
            if ws['U25'].value:
                excel_data.append(f"既存の留意点1: {ws['U25'].value}")
            if ws['C27'].value:
                excel_data.append(f"既存の具体的支援内容2: {ws['C27'].value}")
            if ws['U27'].value:
                excel_data.append(f"既存の留意点2: {ws['U27'].value}")
            if ws['C29'].value:
                excel_data.append(f"既存の具体的支援内容3: {ws['C29'].value}")
            if ws['U29'].value:
                excel_data.append(f"既存の留意点3: {ws['U29'].value}")

            excel_context_for_ai = ""
            if excel_data:
                excel_context_for_ai = "\n\n--- 既存のExcelファイルからの情報 ---\n" + "\n".join(excel_data) + "\n---"
            else:
                excel_context_for_ai = "\n\n--- 既存のExcelファイルからの情報 ---\n既存の計画内容は検出されませんでした。一般的な内容を生成します。\n---"

            # AIに生成を依頼
            response = client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": "あなたは専門的支援実施計画書の作成を支援するAIです。ユーザーから提供された既存のExcelファイルの内容に基づき、以下の形式で「スモールステップ」「長期目標」「短期目標」「具体的支援内容」と、それぞれの「留意点」を**更新**してください。特に、既存の「アセスメント」がある場合は、それを深く考慮して「スモールステップ」を生成してください。既存の情報がある場合は、それを考慮し、より具体的で段階的な（スモールステップな）内容を提案してください。もし既存の情報が不十分な場合は、一般的な内容を生成してください。形式は厳守してください。\n\n「スモールステップ」の内容は、各項目が以下の5領域のいずれかに該当するかを判断し、その領域名を冒頭に`〈領域名〉`の形式で示してください。\n5領域: 「健康・生活」、「運動・感覚」、「認知・行動」、「言語・コミュニケーション」、「人間関係・社会性」\n\n形式：\n① [スモールステップの内容（各項目に〈領域名〉を付与）]\n② [長期目標（1行）]\n[短期目標（1行）]\n③ [具体的支援内容1]\n[留意点1]\n[具体的支援内容2]\n[留意点2]\n[具体的支援内容3]\n[留意点3]\n\n各項目は具体的かつ実践的な内容で記述し、特にスモールステップは段階的に達成可能な内容にしてください。長期目標と短期目標はそれぞれ1行で簡潔に記述し、ラベル（例：長期目標：）は含めないでください。留意点も簡潔に記述してください。"},
                    {"role": "user", "content": f"以下の情報に基づいて計画書の内容を生成または更新してください。\n{excel_context_for_ai}"}
                ],
                temperature=0.7,
            )
            generated_text_from_ai = response.choices[0].message.content
            st.session_state.generated_text = generated_text_from_ai

            # AI生成されたテキストをExcelに書き込む
            part1 = ""
            part2 = ""
            part3 = ""

            if "②" in generated_text_from_ai:
                part1, rest = generated_text_from_ai.split("②", 1)
                part2 = rest
            else:
                part1 = generated_text_from_ai

            if "③" in part2:
                part2, part3 = part2.split("③", 1)

            small_steps = part1.replace("①", "").strip()
            ws['P12'] = small_steps

            # 長期目標と短期目標を明確に分割
            goals_raw = part2.strip().split("\n")
            long_goal = goals_raw[0].strip() if len(goals_raw) >= 1 else ""
            short_goal = goals_raw[1].strip() if len(goals_raw) >= 2 else ""
            
            ws['L19'] = long_goal
            ws['L21'] = short_goal

            # 具体的支援内容と留意点を分割して書き込む
            supports_and_notes = part3.strip().split("\n")
            support_cells = ['C25', 'C27', 'C29']
            note_cells = ['U25', 'U27', 'U29']
            
            for i in range(3):
                support_index = i * 2
                note_index = i * 2 + 1
                
                ws[support_cells[i]] = supports_and_notes[support_index].strip() if support_index < len(supports_and_notes) else ""
                ws[note_cells[i]] = supports_and_notes[note_index].strip() if note_index < len(supports_and_notes) else ""

            # 更新されたExcelファイルを保存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("AIによる更新が完了しました！")
            st.download_button(
                label="更新されたExcelをダウンロード",
                data=output,
                file_name="更新済_専門的支援実施計画書.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.subheader("AIが生成・更新した内容（確認用）")
            st.text_area(
                "",
                value=st.session_state.get("generated_text", ""),
                height=300, 
                disabled=True
            )

        except Exception as e:
            st.error(f"処理中にエラーが発生しました: {e}")
            st.warning("Excelファイルが破損しているか、AI生成に問題が発生した可能性があります。別のファイルをお試しいただくか、しばらく経ってから再度お試しください。")

